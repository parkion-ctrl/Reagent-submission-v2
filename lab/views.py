import io
import json
from urllib.parse import quote_plus

import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group, Permission, User
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.hashers import make_password
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt

from app.core.db import get_connection, get_current_schema
from app.services.history_service import get_history_items, get_history_vendor_options, get_history_reagent_type_options
from app.services.inbound_service import (
    create_bulk_inbound_transactions,
    get_inbound_page_data,
    preview_bulk_inbound_items,
    preview_manual_inbound_items,
)
from app.services.inventory_service import get_inventory_filter_options, get_inventory_items, get_inventory_items_by_code, get_part_equipment_options, get_all_vendor_options, get_disposed_items, cancel_dispose
from app.services.master_service import (
    confirm_bulk_master_items,
    create_master_item,
    delete_master_item,
    get_master_item_by_id,
    get_master_items,
    preview_bulk_master_items_v3,
    update_master_item,
)
from app.services.rawdb_service import (
    get_rawdb_items,
    get_rawdb_item_by_id,
    get_rawdb_filter_options,
    get_master_vendor_equipment_options,
    get_vendor_list,
    get_equipment_list,
    create_vendor,
    update_vendor,
    create_equipment,
    update_equipment,
    create_rawdb_item,
    update_rawdb_item,
    delete_rawdb_item,
    bulk_delete_rawdb_items,
    sync_rawdb_to_inventory,
    upload_rawdb_items_to_master,
)
from app.services.outbound_service import (
    create_bulk_outbound_transactions,
    get_outbound_page_data,
    preview_bulk_outbound_items,
    preview_manual_outbound_items,
)
from app.services.reagent_history_service import (
    dispose_reagent,
    get_reagent_history_filter_options,
    get_reagent_history_items,
    get_old_new_lot_items,
    save_old_new_lot_selection,
    update_opened_at,
    update_parallel_at,
)
from .models import Inventory, TransactionHistory, UserProfile
from app.utils.constants import PART_MAP, get_part_map
from app.services.preset_service import (
    get_presets, get_preset, get_part_item_codes,
    create_preset, update_preset, delete_preset, get_preset_cart_items,
)


def can_access_admin_area(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=["관리자", "개발자"]).exists()


def _is_dlab(request):
    """현재 활성 스키마가 진단검사의학과(dlab)인지 여부."""
    return get_current_schema() == "dlab"


def _get_part(request):
    """파트 파라미터를 반환한다. 미지정 시 로그인 사용자의 기본 파트 사용."""
    part = request.GET.get("part")
    if part is not None:
        return part
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return ""
        try:
            return request.user.profile.part or ""
        except Exception:
            pass
    return ""


@csrf_exempt
def login_page(request):
    error = ""
    next_url = request.POST.get("next") if request.method == "POST" else request.GET.get("next", "")
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            target_url = next_url or "/inventory/"
            if target_url in {"", "/"} or target_url.startswith("/login"):
                target_url = "/inventory/"
            return redirect(target_url)
        try:
            existing = User.objects.get(username=username)
            if not existing.is_active:
                error = "비활성화된 계정입니다. 관리자에게 문의하세요."
            else:
                error = "아이디 또는 비밀번호가 올바르지 않습니다."
        except User.DoesNotExist:
            error = "아이디 또는 비밀번호가 올바르지 않습니다."

    return render(request, "login.html", {"error": error, "next": next_url})


@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def set_dept_view(request):
    """슈퍼유저 전용: 세션의 활성 부서를 전환합니다."""
    if not request.user.is_superuser:
        return redirect(request.GET.get("next", "/inventory/"))
    dept = request.GET.get("dept", "")
    from app.utils.constants import DEPT_SCHEMA_MAP
    if dept in DEPT_SCHEMA_MAP:
        request.session["superuser_active_dept"] = dept
    else:
        request.session.pop("superuser_active_dept", None)
    next_url = request.GET.get("next", "/inventory/")
    return redirect(next_url)


@login_required
@user_passes_test(can_access_admin_area)
def admin_panel(request):
    from datetime import date, timedelta

    one_month_ago = (date.today() - timedelta(days=30)).isoformat()

    # 현재 부서 파악
    if request.user.is_superuser:
        active_dept = request.session.get("superuser_active_dept", "진단검사의학과")
    else:
        profile = UserProfile.objects.filter(user=request.user).first()
        active_dept = profile.department if profile else ""

    user_count = User.objects.filter(profile__department=active_dept).count() if active_dept else User.objects.count()

    # 시약/이력은 get_connection()으로 현재 스키마에서 직접 집계
    inventory_count = 0
    history_count = 0
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM inventory WHERE disposed_at IS NULL")
        inventory_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM transaction_history WHERE tx_date >= %s", (one_month_ago,))
        history_count = cur.fetchone()[0]
        conn.close()
    except Exception:
        pass

    return render(
        request,
        "admin_panel.html",
        {
            "active_menu": "admin_panel",
            "message": request.GET.get("message", ""),
            "stats": {
                "users": user_count,
                "groups": Group.objects.count(),
                "inventory": inventory_count,
                "history": history_count,
            },
        },
    )


@login_required
@user_passes_test(can_access_admin_area)
def admin_users(request):
    from lab.models import DEPARTMENT_CHOICES
    is_superadmin = request.user.is_superuser or request.user.groups.filter(name="개발자").exists()

    my_profile = UserProfile.objects.filter(user=request.user).first()
    my_department = my_profile.department if my_profile else ""

    # 개발자/superuser만 부서 전환 가능, 나머지는 자기 부서 고정
    if is_superadmin:
        selected_department = request.GET.get("department", my_department).strip()
    else:
        selected_department = my_department

    q = request.GET.get("q", "").strip()
    selected_part = request.GET.get("part", "").strip()
    users = User.objects.select_related("profile").all().order_by("username")

    if selected_department:
        users = users.filter(profile__department=selected_department)
    if q:
        users = users.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(profile__employee_no__icontains=q)
        ).distinct().order_by("username")
    if selected_part:
        users = users.filter(profile__part=selected_part)

    profile_map = {
        profile.user_id: profile
        for profile in UserProfile.objects.filter(user_id__in=[user.id for user in users])
    }
    users_data = []
    for user in users:
        group_names = [group.name for group in user.groups.all()]
        if group_names:
            role_labels = group_names
        elif user.is_superuser or user.is_staff:
            role_labels = ["시스템 관리자"]
        else:
            role_labels = ["그룹 없음"]
        profile = profile_map.get(user.id)
        part_code = profile.part if profile else ""
        users_data.append(
            {
                "user": user,
                "employee_no": profile.employee_no if profile else "",
                "part": f"{part_code} ({get_part_map(get_current_schema()).get(part_code, '')})" if part_code else part_code,
                "department": profile.department if profile else "",
                "group_names": group_names,
                "role_labels": role_labels,
            }
        )
    return render(
        request,
        "admin_users.html",
        {
            "active_menu": "admin_panel",
            "users": users_data,
            "q": q,
            "selected_part": selected_part,
            "selected_department": selected_department,
            "part_map": get_part_map(get_current_schema()),
            "department_choices": [c[0] for c in DEPARTMENT_CHOICES],
            "is_superadmin": is_superadmin,
        },
    )


@login_required
@login_required
@user_passes_test(can_access_admin_area)
def admin_user_form(request, user_id=None):
    user_obj = User.objects.filter(id=user_id).first() if user_id else None
    errors = []
    role_group_names = ["개발자", "관리자", "일반", "외부 업체"]
    role_groups = []
    for group_name in role_group_names:
        group_obj, _ = Group.objects.get_or_create(name=group_name)
        role_groups.append(group_obj)

    selected_group_id = 0
    if user_obj:
        current_groups = list(user_obj.groups.order_by("name"))
        if current_groups:
            selected_group_id = current_groups[0].id

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        employee_no = request.POST.get("employee_no", "").strip()
        part = request.POST.get("part", "").strip()
        department = request.POST.get("department", "").strip()
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        password = request.POST.get("password", "")
        is_active = request.POST.get("is_active", "on") == "on"
        selected_group_id = int(request.POST.get("group_id") or 0)

        if not username:
            errors.append("아이디를 입력해 주세요.")
        if user_obj is None and not password:
            errors.append("새 사용자에게는 비밀번호가 필요합니다.")
        is_developer = selected_group_id and Group.objects.filter(id=selected_group_id, name="개발자").exists()
        if not is_developer:
            if not selected_group_id:
                errors.append("권한 그룹을 선택해 주세요.")
            if not department:
                errors.append("부서를 선택해 주세요.")
            if not part:
                errors.append("파트를 선택해 주세요.")

        if not errors:
            target = user_obj or User()
            target.username = username
            target.first_name = first_name
            target.last_name = last_name
            target.is_active = is_active
            if password:
                target.password = make_password(password)
            selected_group = Group.objects.filter(id=selected_group_id).first()
            if selected_group and selected_group.name == "개발자":
                target.is_superuser = True
                target.is_staff = True
            elif selected_group and selected_group.name == "관리자":
                target.is_superuser = False
                target.is_staff = True
            else:
                target.is_superuser = False
                target.is_staff = False
            target.save()
            if selected_group:
                target.groups.set([selected_group])
            else:
                target.groups.clear()
            profile, _ = UserProfile.objects.get_or_create(user=target)
            profile.employee_no = employee_no
            profile.part = part
            profile.department = department
            profile.save()
            return redirect("/admin-users/?message=저장되었습니다.")

    profile_obj = UserProfile.objects.filter(user_id=user_obj.id).first() if user_obj else None
    context = {
        "active_menu": "admin_panel",
        "user_obj": user_obj,
        "profile_employee_no": profile_obj.employee_no if profile_obj else "",
        "profile_part": profile_obj.part if profile_obj else "",
        "profile_department": profile_obj.department if profile_obj else "",
        "groups": role_groups,
        "selected_group_id": selected_group_id,
        "part_map": get_part_map(get_current_schema()),
        "department_choices": ["진단검사의학과", "병리과", "핵의학과", "유해물질"],
        "errors": errors,
    }
    return render(request, "admin_user_form.html", context)


@login_required
@user_passes_test(can_access_admin_area)
@csrf_exempt
def admin_user_delete(request, user_id):
    if request.method != "POST":
        return redirect("admin_users")
    if request.user.id == user_id:
        return redirect("/admin-users/?error=자기 자신은 삭제할 수 없습니다.")
    User.objects.filter(id=user_id).delete()
    return redirect("/admin-users/?message=삭제되었습니다.")


@login_required
@user_passes_test(can_access_admin_area)
def admin_groups(request):
    q = request.GET.get("q", "").strip()
    groups = Group.objects.all().order_by("name")
    if q:
        groups = groups.filter(name__icontains=q)
    return render(
        request,
        "admin_groups.html",
        {
            "active_menu": "admin_panel",
            "groups": groups,
            "q": q,
        },
    )


@login_required
@user_passes_test(can_access_admin_area)
def admin_group_form(request, group_id=None):
    group_obj = Group.objects.filter(id=group_id).first() if group_id else None
    permissions = Permission.objects.select_related("content_type").order_by("content_type__app_label", "content_type__model", "codename")
    errors = []
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        permission_ids = [int(v) for v in request.POST.getlist("permissions") if str(v).strip()]
        selected_permission_ids = permission_ids
        if not name:
            errors.append("그룹명을 입력해 주세요.")
        if not errors:
            target = group_obj or Group()
            target.name = name
            target.save()
            target.permissions.set(Permission.objects.filter(id__in=permission_ids))
            return redirect("/admin-groups/?message=저장되었습니다.")
    return render(
        request,
        "admin_group_form.html",
        {
            "active_menu": "admin_panel",
            "group_obj": group_obj,
            "permissions": permissions,
            "selected_permission_ids": selected_permission_ids,
            "errors": errors,
        },
    )


@login_required
@user_passes_test(can_access_admin_area)
@csrf_exempt
def admin_group_delete(request, group_id):
    if request.method != "POST":
        return redirect("admin_groups")
    Group.objects.filter(id=group_id).delete()
    return redirect("/admin-groups/?message=삭제되었습니다.")


@login_required
@user_passes_test(can_access_admin_area)
def admin_parts(request):
    from lab.models import Part
    from app.core.db import ALL_SCHEMAS
    import psycopg
    from app.core.db import PG_HOST, PG_PORT, PG_DBNAME, PG_USER, PG_PASSWORD

    message = request.GET.get("message", "")
    current_schema = get_current_schema()
    # dlab은 schema_name=None, 나머지는 해당 schema_name으로 필터
    if current_schema == "dlab" or current_schema is None:
        qs = Part.objects.filter(schema_name__isnull=True)
    else:
        qs = Part.objects.filter(schema_name=current_schema)
    parts = list(qs)

    # 현재 부서 스키마 기준 파트별 사용 건수 집계
    usage = {}
    try:
        schemas = ALL_SCHEMAS if (current_schema == "dlab" or current_schema is None) else ([current_schema] if current_schema else [])
        conn = psycopg.connect(host=PG_HOST, port=PG_PORT, dbname=PG_DBNAME,
                               user=PG_USER, password=PG_PASSWORD)
        cur = conn.cursor()
        for schema in schemas:
            cur.execute(f"SELECT part, COUNT(*) FROM {schema}.inventory WHERE disposed_at IS NULL GROUP BY part")
            for row in cur.fetchall():
                code = (row[0] or "").strip()
                usage[code] = usage.get(code, 0) + int(row[1])
        conn.close()
    except Exception:
        pass

    parts_data = [{"obj": p, "usage": usage.get(p.code, 0)} for p in parts]

    return render(request, "admin_parts.html", {
        "active_menu": "admin_panel",
        "parts_data": parts_data,
        "message": message,
    })


@login_required
@user_passes_test(can_access_admin_area)
def admin_part_form(request, part_id=None):
    from lab.models import Part
    current_schema = get_current_schema()
    # dlab은 schema_name=None, 나머지는 해당 schema_name
    save_schema = None if (current_schema == "dlab" or current_schema is None) else current_schema

    part_obj = Part.objects.filter(id=part_id).first() if part_id else None
    errors = []

    if request.method == "POST":
        code = request.POST.get("code", "").strip().upper()
        name = request.POST.get("name", "").strip()

        if not code:
            errors.append("파트 코드는 필수입니다.")
        if not name:
            errors.append("파트명은 필수입니다.")
        if code and Part.objects.filter(code=code, schema_name=save_schema).exclude(id=part_id).exists():
            errors.append(f"파트 코드 '{code}'는 이미 존재합니다.")

        if not errors:
            if part_obj:
                part_obj.code = code
                part_obj.name = name
                part_obj.save()
            else:
                Part.objects.create(code=code, name=name, schema_name=save_schema)
            return redirect("/admin-parts/?message=저장되었습니다.")

    return render(request, "admin_part_form.html", {
        "active_menu": "admin_panel",
        "part_obj": part_obj,
        "errors": errors,
    })


@login_required
@user_passes_test(can_access_admin_area)
@csrf_exempt
def admin_part_delete(request, part_id):
    if request.method != "POST":
        return redirect("admin_parts")
    from lab.models import Part
    Part.objects.filter(id=part_id).delete()
    return redirect("/admin-parts/?message=삭제되었습니다.")


@login_required
def root_redirect(request):
    return redirect("login")


def health_check(request):
    return JsonResponse({"status": "ok"})


def get_master_base_context():
    return {
        "active_menu": "master",
        "items": get_master_items(),
        "part_map": get_part_map(get_current_schema()),
        "part": "",
        "q": "",
        "sort": "",
        "order": "",
        "show_form": "",
        "message": "",
        "error": "",
        "reagent_type": "",
        "edit_item": None,
        "equipment": "",
        "vendor": "",
        "hazardous": "",
        "filter_options": get_inventory_filter_options(),
    }


@login_required
def inventory_page(request):
    part = _get_part(request)
    q = request.GET.get("q", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    hazardous_grade = request.GET.get("hazardous_grade", "")
    is_new_lot = request.GET.get("is_new_lot", "")
    expiry_filter = request.GET.get("expiry_filter", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    view_mode = request.GET.get("view_mode", "lot")
    if view_mode not in ("lot", "code"):
        view_mode = "lot"

    if view_mode == "code":
        items = get_inventory_items_by_code(
            part=part, q=q, reagent_type=reagent_type,
            equipment=equipment, vendor=vendor,
            hazardous=hazardous, hazardous_grade=hazardous_grade,
            sort=sort, order=order,
        )
    else:
        items = get_inventory_items(
            part=part, q=q, reagent_type=reagent_type,
            equipment=equipment, vendor=vendor,
            hazardous=hazardous, hazardous_grade=hazardous_grade,
            is_new_lot=is_new_lot, expiry_filter=expiry_filter,
            sort=sort, order=order,
        )
    context = {
        "active_menu": "inventory",
        "items": items,
        "part_map": get_part_map(get_current_schema()),
        "part": part,
        "q": q,
        "reagent_type": reagent_type,
        "equipment": equipment,
        "vendor": vendor,
        "hazardous": hazardous,
        "hazardous_grade": hazardous_grade,
        "is_new_lot": is_new_lot,
        "expiry_filter": expiry_filter,
        "sort": sort,
        "order": order,
        "view_mode": view_mode,
        "filter_options": get_inventory_filter_options(),
        "filtered_equipments": get_part_equipment_options(part),
        "filtered_vendors": get_all_vendor_options(),
    }
    return render(request, "inventory.html", context)


@login_required
def inventory_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    part = request.GET.get("part", "")
    q = request.GET.get("q", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    expiry_filter = request.GET.get("expiry_filter", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")

    items = get_inventory_items(
        part=part, q=q, reagent_type=reagent_type, equipment=equipment,
        vendor=vendor, hazardous=hazardous, expiry_filter=expiry_filter,
        sort=sort, order=order,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시약재고"

    headers = ["파트", "품목코드", "품목명", "Lot No", "유효기간", "규격", "단위",
               "시약종류", "장비", "업체", "현재재고", "안전재고", "필요수량", "유해화학물질"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item["part_label"], item["item_code"], item["item_name"],
            item["lot_no"] or "", item["expiry_date"] or "", item["spec"] or "",
            item["unit"] or "", item["reagent_type"] or "", item["equipment"] or "",
            item["vendor"] or "", item["current_stock"], item["safety_stock"],
            item["required_qty"], item["hazardous"] or "",
        ])

    col_widths = [16, 14, 40, 16, 12, 16, 8, 14, 18, 18, 10, 10, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    encoded = quote("시약재고.xlsx")
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@login_required
def master_page(request):
    part = _get_part(request)
    q = request.GET.get("q", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    show_form = request.GET.get("show_form", "")
    edit_id = request.GET.get("edit_id", "")
    message = request.GET.get("message", "")
    error = request.GET.get("error", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    hazardous_grade = request.GET.get("hazardous_grade", "")
    is_new_lot = request.GET.get("is_new_lot", "")

    from collections import defaultdict
    equip_by_part = defaultdict(list)
    for e in get_equipment_list():
        equip_by_part[e["part"]].append(e["name"])

    items = get_master_items(
        part=part,
        q=q,
        reagent_type=reagent_type,
        equipment=equipment,
        vendor=vendor,
        hazardous=hazardous,
        hazardous_grade=hazardous_grade,
        is_new_lot=is_new_lot,
        sort=sort,
        order=order,
    )
    context = {
        "active_menu": "master",
        "items": items,
        "part_map": get_part_map(get_current_schema()),
        "part": part,
        "q": q,
        "sort": sort,
        "order": order,
        "show_form": show_form,
        "message": message,
        "error": error,
        "reagent_type": reagent_type,
        "edit_item": get_master_item_by_id(int(edit_id)) if edit_id else None,
        "equipment": equipment,
        "vendor": vendor,
        "hazardous": hazardous,
        "hazardous_grade": hazardous_grade,
        "is_new_lot": is_new_lot,
        "filter_options": get_inventory_filter_options(),
        "filtered_equipments": get_part_equipment_options(part),
        "filtered_vendors": get_all_vendor_options(),
        "vendor_options": [v["name"] for v in get_vendor_list()],
        "equipment_by_part_json": json.dumps(dict(equip_by_part)),
        "current_schema": get_current_schema(),
    }
    return render(request, "master.html", context)


@login_required
def master_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    part = request.GET.get("part", "")
    q = request.GET.get("q", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")

    items = get_master_items(
        part=part, q=q, reagent_type=reagent_type, equipment=equipment,
        vendor=vendor, hazardous=hazardous, sort=sort, order=order,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시약마스터"

    is_path = get_current_schema() == "path"
    if is_path:
        headers = ["파트", "품목코드", "품목명", "CAS-NO", "Lot No", "유효기간", "규격", "단위",
                   "시약종류", "장비", "업체", "현재재고", "안전재고", "유해화학물질"]
    else:
        headers = ["파트", "품목코드", "품목명", "Lot No", "유효기간", "규격", "단위",
                   "시약종류", "장비", "업체", "현재재고", "안전재고", "유해화학물질"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        if is_path:
            ws.append([
                item.get("part_label", ""), item.get("item_code", ""), item.get("item_name", ""),
                item.get("cas_no") or "",
                item.get("lot_no") or "", item.get("expiry_date") or "", item.get("spec") or "",
                item.get("unit") or "", item.get("reagent_type") or "", item.get("equipment") or "",
                item.get("vendor") or "", item.get("current_stock", 0), item.get("safety_stock", 0),
                item.get("hazardous") or "",
            ])
        else:
            ws.append([
                item.get("part_label", ""), item.get("item_code", ""), item.get("item_name", ""),
                item.get("lot_no") or "", item.get("expiry_date") or "", item.get("spec") or "",
                item.get("unit") or "", item.get("reagent_type") or "", item.get("equipment") or "",
                item.get("vendor") or "", item.get("current_stock", 0), item.get("safety_stock", 0),
                item.get("hazardous") or "",
            ])

    col_widths = [16, 14, 40, 16, 12, 16, 8, 14, 18, 18, 10, 10, 14] if not is_path else [16, 14, 40, 18, 16, 12, 16, 8, 14, 18, 18, 10, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    encoded = quote("시약마스터.xlsx")
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@csrf_exempt
@login_required
def master_create(request):
    if request.method != "POST":
        return redirect("master")

    ok, msg = create_master_item(
        hazardous=request.POST.get("hazardous", ""),
        hazardous_grade=request.POST.get("hazardous_grade", ""),
        part=request.POST.get("part", ""),
        item_code=request.POST.get("item_code", ""),
        item_name=request.POST.get("item_name", ""),
        lot_no=request.POST.get("lot_no", ""),
        expiry_date=request.POST.get("expiry_date", ""),
        spec=request.POST.get("spec", ""),
        unit=request.POST.get("unit", ""),
        reagent_type=request.POST.get("reagent_type", ""),
        equipment=request.POST.get("equipment", ""),
        vendor=request.POST.get("vendor", ""),
        safety_stock=int(request.POST.get("safety_stock", 0) or 0),
        cas_no=request.POST.get("cas_no", ""),
    )
    if ok:
        return redirect("/master/?message=항목이 등록되었습니다.")
    return redirect(f"/master/?show_form=1&error={quote_plus(msg)}")


@csrf_exempt
@login_required
def master_edit_submit(request, item_id: int):
    if request.method != "POST":
        return redirect("master")

    update_master_item(
        item_id=item_id,
        hazardous=request.POST.get("hazardous", ""),
        hazardous_grade=request.POST.get("hazardous_grade", ""),
        part=request.POST.get("part", ""),
        item_code=request.POST.get("item_code", ""),
        item_name=request.POST.get("item_name", ""),
        lot_no=request.POST.get("lot_no", ""),
        expiry_date=request.POST.get("expiry_date", ""),
        spec=request.POST.get("spec", ""),
        unit=request.POST.get("unit", ""),
        reagent_type=request.POST.get("reagent_type", ""),
        equipment=request.POST.get("equipment", ""),
        vendor=request.POST.get("vendor", ""),
        safety_stock=int(request.POST.get("safety_stock", 0) or 0),
        cas_no=request.POST.get("cas_no", ""),
    )
    return redirect("/master/?message=항목이 수정되었습니다.")


@csrf_exempt
@login_required
def master_delete(request, item_id: int):
    if request.method != "POST":
        return redirect("master")
    delete_master_item(item_id)
    return redirect("/master/?message=항목이 삭제되었습니다.")


@csrf_exempt
@login_required
def master_dispose(request, item_id: int):
    if request.method != "POST":
        return redirect("master")
    dispose_reagent(item_id=item_id, reason="자동 폐기", disposal_type="MANUAL")
    return redirect("/master/?message=시약이 폐기되었습니다.")


@csrf_exempt
@login_required
def master_bulk_delete(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "잘못된 요청입니다."}, status=405)
    from django.contrib.auth import authenticate
    password = request.POST.get("confirm_password", "")
    user = authenticate(request, username=request.user.username, password=password)
    if user is None:
        return JsonResponse({"ok": False, "error": "비밀번호가 올바르지 않습니다."})
    item_ids = [int(item_id) for item_id in request.POST.getlist("item_ids") if str(item_id).strip()]
    deleted, skipped = [], []
    for item_id in item_ids:
        try:
            delete_master_item(item_id)
            deleted.append(item_id)
        except Exception as e:
            if "foreign key" in str(e).lower() or "참조키" in str(e) or "ForeignKey" in type(e).__name__:
                skipped.append(item_id)
            else:
                raise
    parts = []
    if deleted:
        parts.append(f"{len(deleted)}개 항목 삭제 완료")
    if skipped:
        parts.append(f"{len(skipped)}개 항목은 입출고 이력이 있어 삭제할 수 없습니다. 삭제 대신 폐기 처리를 진행해주세요.")
    if skipped and not deleted:
        return JsonResponse({"ok": False, "error": parts[0]})
    return JsonResponse({"ok": True, "message": " / ".join(parts)})


@csrf_exempt
@login_required
def master_bulk_dispose(request):
    if request.method != "POST":
        return redirect("master")
    item_ids = [int(item_id) for item_id in request.POST.getlist("item_ids") if str(item_id).strip()]
    for item_id in item_ids:
        dispose_reagent(item_id=item_id, reason="자동 폐기", disposal_type="MANUAL")
    return redirect(f"/master/?message={len(item_ids)}개 항목이 폐기되었습니다.")


@login_required
def download_master_upload_template(request):
    df = pd.DataFrame(
        [
            {
                "hazardous": "Y",
                "hazardous_grade": "1",
                "part": "TA",
                "item_code": "CRP001",
                "item_name": "CRP 시약",
                "lot_no": "",
                "expiry_date": "20261231",
                "spec": "500mL",
                "unit": "EA",
                "reagent_type": "1",
                "equipment": "c702",
                "vendor": "Roche",
                "safety_stock": 10,
            }
        ]
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="master_upload_template")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=master_upload_template.xlsx"
    return response


@csrf_exempt
@login_required
def upload_master_preview(request):
    if request.method != "POST":
        return redirect("master")
    try:
        file = request.FILES["file"]
        filename = file.name.lower()
        if filename.endswith(".xlsx"):
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                df = pd.read_excel(file, dtype=str)
        else:
            df = pd.read_csv(file, dtype=str)
        df = df.fillna("")
        preview_result = preview_bulk_master_items_v3(df)
        context = get_master_base_context()
        context.update(
            {
                "preview_mode": True,
                "preview_rows": preview_result["preview_rows"],
                "preview_json": json.dumps(preview_result["upload_rows"], ensure_ascii=False),
                "total_count": preview_result["total_count"],
                "valid_count": preview_result["valid_count"],
                "duplicate_count": preview_result["duplicate_count"],
                "duplicate_names": preview_result["duplicate_names"],
                "invalid_count": preview_result["invalid_count"],
                "invalid_messages": preview_result["invalid_messages"],
            }
        )
        return render(request, "master.html", context)
    except Exception as exc:
        context = get_master_base_context()
        context["upload_error"] = str(exc)
        return render(request, "master.html", context)


@csrf_exempt
@login_required
def upload_master_confirm(request):
    if request.method != "POST":
        return redirect("master")
    try:
        rows = json.loads(request.POST.get("upload_data", "[]"))
        result = confirm_bulk_master_items(rows)
        context = get_master_base_context()
        context.update(
            {
                "upload_done": True,
                "total_uploaded": result["total"],
                "success_count": result["success"],
                "fail_count": result["fail"],
                "fail_messages": result["fail_messages"],
            }
        )
        return render(request, "master.html", context)
    except Exception as exc:
        context = get_master_base_context()
        context["upload_error"] = str(exc)
        return render(request, "master.html", context)


@login_required
def inbound_page(request):
    q = request.GET.get("q", "")
    part = _get_part(request)
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    equipment = request.GET.get("equipment", "")
    reagent_type = request.GET.get("reagent_type", "")
    context = {
        "active_menu": "inbound",
        "tx_mode": "inbound",
        "tx_title": "시약 입고 등록",
        "tx_button_label": "입고 등록",
        "tx_qty_label": "입고수량",
        "tx_date_label": "입고일자",
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
        "q": q,
        "part": part,
        "sort": sort,
        "order": order,
        "equipment": equipment,
        "reagent_type": reagent_type,
        "part_map": get_part_map(get_current_schema()),
    }
    context.update(get_inbound_page_data(q=q, part=part, sort=sort, order=order, equipment=equipment, reagent_type=reagent_type))
    context["presets"] = get_presets(part=part)
    return render(request, "transaction_entry.html", context)


def get_inbound_base_context(q: str = "", part: str = "", sort: str = "", order: str = "", equipment: str = "", reagent_type: str = ""):
    context = {
        "active_menu": "inbound",
        "tx_mode": "inbound",
        "tx_title": "시약 입고 등록",
        "tx_button_label": "입고 등록",
        "tx_qty_label": "입고수량",
        "tx_date_label": "입고일자",
        "message": "",
        "error": "",
        "q": q,
        "part": part,
        "sort": sort,
        "order": order,
        "equipment": equipment,
        "reagent_type": reagent_type,
        "part_map": get_part_map(get_current_schema()),
    }
    context.update(get_inbound_page_data(q=q, part=part, sort=sort, order=order, equipment=equipment, reagent_type=reagent_type))
    return context


@csrf_exempt
@login_required
def inbound_bulk_create(request):
    if request.method != "POST":
        return redirect("inbound")
    try:
        rows = json.loads(request.POST.get("rows_json", "[]"))
    except json.JSONDecodeError:
        return redirect(f"/inbound/?error={quote_plus('선택된 항목 데이터를 읽을 수 없습니다.')}")

    created_by = request.user.get_full_name() or request.user.username
    created_by_empno = getattr(request.user, "profile", None) and request.user.profile.employee_no or ""
    ok, msg = create_bulk_inbound_transactions(rows, created_by=created_by, created_by_empno=created_by_empno)
    target = "message" if ok else "error"
    return redirect(f"/inbound/?{target}={quote_plus(msg)}")


@csrf_exempt
@login_required
def inbound_bulk_preview(request):
    if request.method != "POST":
        return redirect("inbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        rows = json.loads(request.POST.get("rows_json", "[]"))
        preview_result = preview_manual_inbound_items(rows)
        context = get_inbound_base_context(q=q, part=part)
        context.update(
            {
                "preview_mode": True,
                "preview_rows": preview_result["preview_rows"],
                "preview_json": json.dumps(preview_result["upload_rows"], ensure_ascii=False),
                "total_count": preview_result["total_count"],
                "valid_count": preview_result["valid_count"],
                "invalid_count": preview_result["invalid_count"],
                "invalid_messages": preview_result["invalid_messages"],
                "new_lot_messages": preview_result.get("new_lot_messages", []),
            }
        )
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_inbound_base_context(q=q, part=part)
        context["error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@login_required
def download_inbound_upload_template(request):
    df = pd.DataFrame(
        [{"item_code": "CRP001", "lot_no": "LOT202603", "expiry_date": "20271231", "qty": 5, "tx_date": "20260409"}]
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="inbound_upload_template")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=inbound_upload_template.xlsx"
    return response


@csrf_exempt
@login_required
def inbound_upload_preview(request):
    if request.method != "POST":
        return redirect("inbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        file = request.FILES["file"]
        filename = file.name.lower()
        if filename.endswith(".xlsx"):
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                df = pd.read_excel(file, dtype=str)
        else:
            df = pd.read_csv(file, dtype=str)
        df = df.fillna("")
        preview_result = preview_bulk_inbound_items(df)
        new_lot_messages = preview_result.get("new_lot_messages", [])
        context = get_inbound_base_context(q=q, part=part)
        context.update(
            {
                "preview_mode": True,
                "preview_rows": preview_result["preview_rows"],
                "preview_json": json.dumps(preview_result["upload_rows"], ensure_ascii=False),
                "total_count": preview_result["total_count"],
                "valid_count": preview_result["valid_count"],
                "invalid_count": preview_result["invalid_count"],
                "invalid_messages": preview_result["invalid_messages"],
                "new_lot_messages": new_lot_messages,
            }
        )
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_inbound_base_context(q=q, part=part)
        context["upload_error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@csrf_exempt
@login_required
def inbound_upload_confirm(request):
    if request.method != "POST":
        return redirect("inbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        rows = json.loads(request.POST.get("upload_data", "[]"))
        created_by = request.user.get_full_name() or request.user.username
        created_by_empno = getattr(request.user, "profile", None) and request.user.profile.employee_no or ""
        ok, msg = create_bulk_inbound_transactions(rows, created_by=created_by, created_by_empno=created_by_empno)
        context = get_inbound_base_context(q=q, part=part)
        if ok:
            context["message"] = msg
        else:
            context["error"] = msg
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_inbound_base_context(q=q, part=part)
        context["upload_error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@login_required
def outbound_page(request):
    q = request.GET.get("q", "")
    part = _get_part(request)
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    equipment = request.GET.get("equipment", "")
    reagent_type = request.GET.get("reagent_type", "")
    context = {
        "active_menu": "outbound",
        "tx_mode": "outbound",
        "tx_title": "시약 출고 등록",
        "tx_button_label": "출고 등록",
        "tx_qty_label": "출고수량",
        "tx_date_label": "출고일자",
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
        "q": q,
        "part": part,
        "sort": sort,
        "order": order,
        "equipment": equipment,
        "reagent_type": reagent_type,
        "part_map": get_part_map(get_current_schema()),
    }
    context.update(get_outbound_page_data(q=q, part=part, sort=sort, order=order, equipment=equipment, reagent_type=reagent_type))
    return render(request, "transaction_entry.html", context)


def get_outbound_base_context(q: str = "", part: str = "", sort: str = "", order: str = "", equipment: str = "", reagent_type: str = ""):
    context = {
        "active_menu": "outbound",
        "tx_mode": "outbound",
        "tx_title": "시약 출고 등록",
        "tx_button_label": "출고 등록",
        "tx_qty_label": "출고수량",
        "tx_date_label": "출고일자",
        "message": "",
        "error": "",
        "q": q,
        "part": part,
        "sort": sort,
        "order": order,
        "equipment": equipment,
        "reagent_type": reagent_type,
        "part_map": get_part_map(get_current_schema()),
    }
    context.update(get_outbound_page_data(q=q, part=part, sort=sort, order=order, equipment=equipment, reagent_type=reagent_type))
    return context


@csrf_exempt
@login_required
def outbound_bulk_create(request):
    if request.method != "POST":
        return redirect("outbound")
    try:
        rows = json.loads(request.POST.get("rows_json", "[]"))
    except json.JSONDecodeError:
        return redirect(f"/outbound/?error={quote_plus('선택된 항목 데이터를 읽을 수 없습니다.')}")

    created_by = request.user.get_full_name() or request.user.username
    created_by_empno = getattr(request.user, "profile", None) and request.user.profile.employee_no or ""
    ok, msg = create_bulk_outbound_transactions(rows, created_by=created_by, created_by_empno=created_by_empno)
    target = "message" if ok else "error"
    return redirect(f"/outbound/?{target}={quote_plus(msg)}")


@csrf_exempt
@login_required
def outbound_bulk_preview(request):
    if request.method != "POST":
        return redirect("outbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        rows = json.loads(request.POST.get("rows_json", "[]"))
        preview_result = preview_manual_outbound_items(rows)
        context = get_outbound_base_context(q=q, part=part)
        context.update(
            {
                "preview_mode": True,
                "preview_rows": preview_result["preview_rows"],
                "preview_json": json.dumps(preview_result["upload_rows"], ensure_ascii=False),
                "total_count": preview_result["total_count"],
                "valid_count": preview_result["valid_count"],
                "invalid_count": preview_result["invalid_count"],
                "invalid_messages": preview_result["invalid_messages"],
            }
        )
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_outbound_base_context(q=q, part=part)
        context["error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@login_required
def download_outbound_upload_template(request):
    df = pd.DataFrame(
        [{"item_code": "CRP001", "lot_no": "LOT202603", "expiry_date": "20271231", "qty": 2, "tx_date": "20260409"}]
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="outbound_upload_template")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=outbound_upload_template.xlsx"
    return response


@csrf_exempt
@login_required
def outbound_upload_preview(request):
    if request.method != "POST":
        return redirect("outbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        file = request.FILES["file"]
        filename = file.name.lower()
        if filename.endswith(".xlsx"):
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                df = pd.read_excel(file, dtype=str)
        else:
            df = pd.read_csv(file, dtype=str)
        df = df.fillna("")
        preview_result = preview_bulk_outbound_items(df)
        context = get_outbound_base_context(q=q, part=part)
        context.update(
            {
                "preview_mode": True,
                "preview_rows": preview_result["preview_rows"],
                "preview_json": json.dumps(preview_result["upload_rows"], ensure_ascii=False),
                "total_count": preview_result["total_count"],
                "valid_count": preview_result["valid_count"],
                "invalid_count": preview_result["invalid_count"],
                "invalid_messages": preview_result["invalid_messages"],
            }
        )
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_outbound_base_context(q=q, part=part)
        context["upload_error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@csrf_exempt
@login_required
def outbound_upload_confirm(request):
    if request.method != "POST":
        return redirect("outbound")
    q = request.POST.get("q", "")
    part = request.POST.get("part", "")
    try:
        rows = json.loads(request.POST.get("upload_data", "[]"))
        created_by = request.user.get_full_name() or request.user.username
        created_by_empno = getattr(request.user, "profile", None) and request.user.profile.employee_no or ""
        ok, msg = create_bulk_outbound_transactions(rows, created_by=created_by, created_by_empno=created_by_empno)
        context = get_outbound_base_context(q=q, part=part)
        if ok:
            context["message"] = msg
        else:
            context["error"] = msg
        return render(request, "transaction_entry.html", context)
    except Exception as exc:
        context = get_outbound_base_context(q=q, part=part)
        context["upload_error"] = str(exc)
        return render(request, "transaction_entry.html", context)


@login_required
def history_page(request):
    from datetime import date, timedelta

    tx_type = request.GET.get("tx_type", "")
    part = _get_part(request)
    q = request.GET.get("q", "")
    vendor = request.GET.get("vendor", "")
    reagent_type = request.GET.get("reagent_type", "")
    disposed = request.GET.get("disposed", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    period = request.GET.get("period", "1d")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "asc")
    today = date.today()

    if not date_from and not date_to:
        if period == "7d":
            date_from = (today - timedelta(days=7)).isoformat()
        elif period == "1m":
            date_from = (today - timedelta(days=30)).isoformat()
        elif period == "6m":
            date_from = (today - timedelta(days=183)).isoformat()
        else:
            period = "1d"
            date_from = today.isoformat()
        date_to = today.isoformat()

    items = get_history_items(
        tx_type=tx_type,
        part=part,
        q=q,
        vendor=vendor,
        reagent_type=reagent_type,
        disposed=disposed,
        date_from=date_from,
        date_to=date_to,
        sort=sort,
        order=order,
    )
    context = {
        "active_menu": "history",
        "items": items,
        "tx_type": tx_type,
        "part": part,
        "q": q,
        "vendor": vendor,
        "reagent_type": reagent_type,
        "disposed": disposed,
        "date_from": date_from,
        "date_to": date_to,
        "period": period,
        "sort": sort,
        "order": order,
        "part_map": get_part_map(get_current_schema()),
        "vendor_options": get_all_vendor_options(),
        "reagent_type_options": get_history_reagent_type_options(),
    }
    return render(request, "history.html", context)


@login_required
def history_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from urllib.parse import quote
    from datetime import date, timedelta

    tx_type = request.GET.get("tx_type", "")
    part = request.GET.get("part", "") or _get_part(request)
    q = request.GET.get("q", "")
    vendor = request.GET.get("vendor", "")
    reagent_type = request.GET.get("reagent_type", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    period = request.GET.get("period", "1d")
    today = date.today()

    if not date_from and not date_to:
        if period == "1d":
            date_from = today.isoformat()
        elif period == "1m":
            date_from = (today - timedelta(days=30)).isoformat()
        elif period == "6m":
            date_from = (today - timedelta(days=183)).isoformat()
        else:
            date_from = (today - timedelta(days=7)).isoformat()
        date_to = today.isoformat()

    items = get_history_items(tx_type=tx_type, part=part, q=q, vendor=vendor, reagent_type=reagent_type, date_from=date_from, date_to=date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "입출고이력"

    headers = ["거래일", "구분", "파트", "품목코드", "품목명", "Lot No", "업체", "시약구분", "입고수량", "출고수량", "잔여재고", "등록자"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        if item.get("billing_label"):
            label = item["billing_label"]
        else:
            label = item["tx_type_label"]
        ws.append([
            item["tx_date"],
            label,
            item["part_label"],
            item["item_code"],
            item["item_name"],
            item["lot_no"] or "",
            item.get("vendor") or "",
            item.get("reagent_type_label") or "",
            item["inbound_qty"] or "",
            item["outbound_qty"] or "",
            item["remaining_stock"],
            item["created_by"] or "",
        ])

    col_widths = [12, 12, 18, 14, 40, 16, 18, 12, 10, 10, 10, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"입출고이력_{date_from}_{date_to}.xlsx"
    encoded = quote(filename)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@login_required
def history_billing_type(request):
    """선택 이력의 billing_type 업데이트 API"""
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
    from django.http import JsonResponse
    try:
        data = json.loads(request.body)
        ids = [int(i) for i in data.get("ids", [])]
        new_billing_type = data.get("billing_type")  # None, 'FREE_IN', 'PROV_OUT'
        if not ids:
            return JsonResponse({"ok": False, "error": "ids required"})

        from app.services.history_service import recalculate_current_stock

        conn = get_connection()
        cursor = conn.cursor()
        placeholders = ",".join(["%s"] * len(ids))

        cursor.execute(
            f"UPDATE transaction_history SET billing_type = %s WHERE id IN ({placeholders})",
            [new_billing_type] + ids,
        )
        conn.commit()
        conn.close()

        # billing_type 변경 후 전체 재고 재계산
        recalculate_current_stock()

        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})


@login_required
def billing_page(request):
    from app.services.billing_service import get_billing_items
    from datetime import date, timedelta

    today = date.today()
    part = _get_part(request)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "asc")

    if not date_from and not date_to:
        date_from = (today - timedelta(days=30)).isoformat()
        date_to = today.isoformat()

    items = get_billing_items(part=part, date_from=date_from, date_to=date_to)

    if sort == "item_code":
        items.sort(key=lambda x: (x.get("item_code") or "").lower(), reverse=(order == "desc"))
    elif sort == "item_name":
        items.sort(key=lambda x: (x.get("item_name") or "").lower(), reverse=(order == "desc"))

    context = {
        "active_menu": "billing",
        "items": items,
        "date_from": date_from,
        "date_to": date_to,
        "part": part,
        "sort": sort,
        "order": order,
        "part_map": get_part_map(get_current_schema()),
    }
    return render(request, "billing.html", context)


@login_required
def billing_export(request):
    from app.services.billing_service import get_billing_items
    from datetime import date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from urllib.parse import quote

    today = date.today()
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    part = request.GET.get("part", "")

    items = get_billing_items(part=part, date_from=date_from, date_to=date_to)
    out_items = [i for i in items if i["tx_type"] == "OUT"]
    in_items  = [i for i in items if i["tx_type"] == "IN"]

    wb = openpyxl.Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center", vertical="center")

    OUT_HEADERS = ["구분", "파트", "품목코드", "품목명", "업체", "수량", "단가", "비용", "비고"]
    IN_HEADERS  = ["구분", "파트", "품목코드", "품목명", "업체", "수량", "단가", "비용", "비고"]

    def write_out_sheet(ws, rows):
        ws.append(OUT_HEADERS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        total_cost = 0
        for row in rows:
            up = row.get("unit_price")
            cost = row.get("cost")
            ws.append([
                row["tx_type_label"],
                row["part_label"],
                row["item_code"],
                row["item_name"],
                row["vendor"] or "",
                row["qty_display"],
                int(up) if up is not None else "",
                cost if cost is not None else "",
                row["note"] or "",
            ])
            if cost: total_cost += cost

        # 총 정산액 행
        if rows:
            total_row = [""] * 9
            total_row[6] = "총 정산액"
            total_row[7] = total_cost
            ws.append(total_row)
            last = ws.max_row
            for cell in ws[last]:
                cell.font = Font(bold=True)

        col_widths = [10, 18, 16, 40, 18, 8, 12, 14, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    def write_in_sheet(ws, rows):
        ws.append(IN_HEADERS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        total_cost = 0
        for row in rows:
            up = row.get("unit_price")
            cost = row.get("cost")
            ws.append([
                row["tx_type_label"],
                row["part_label"],
                row["item_code"],
                row["item_name"],
                row["vendor"] or "",
                row["qty_display"],
                int(up) if up is not None else "",
                cost if cost is not None else "",
                row["note"] or "",
            ])
            if cost: total_cost += cost

        if rows:
            total_row = [""] * 9
            total_row[6] = "총 정산액"
            total_row[7] = total_cost
            ws.append(total_row)
            last = ws.max_row
            for cell in ws[last]:
                cell.font = Font(bold=True)

        col_widths = [10, 18, 16, 40, 18, 8, 12, 14, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 출고 시트
    ws_out = wb.active
    ws_out.title = "출고"
    write_out_sheet(ws_out, out_items)

    # 입고 시트
    ws_in = wb.create_sheet("입고")
    write_in_sheet(ws_in, in_items)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"시약청구_{date_from}_{date_to}.xlsx"
    encoded = quote(filename)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@login_required
@user_passes_test(can_access_admin_area)
def history_admin_page(request):
    from datetime import date, timedelta

    tx_type = request.GET.get("tx_type", "")
    part = _get_part(request)
    q = request.GET.get("q", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    period = request.GET.get("period", "1d")
    today = date.today()

    if not date_from and not date_to:
        if period == "7d":
            date_from = (today - timedelta(days=7)).isoformat()
        elif period == "1m":
            date_from = (today - timedelta(days=30)).isoformat()
        elif period == "6m":
            date_from = (today - timedelta(days=183)).isoformat()
        else:
            period = "1d"
            date_from = today.isoformat()
        date_to = today.isoformat()

    context = {
        "active_menu": "admin_panel",
        "items": get_history_items(tx_type=tx_type, part=part, q=q, date_from=date_from, date_to=date_to),
        "tx_type": tx_type,
        "part": part,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "period": period,
        "part_map": get_part_map(get_current_schema()),
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
    }
    return render(request, "history_admin.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def history_admin_edit(request, record_id):
    if request.method != "POST":
        return redirect("history_admin")
    tx_type = request.POST.get("tx_type", "").strip()
    tx_date = request.POST.get("tx_date", "").strip()
    qty = int(request.POST.get("qty", 0) or 0)
    referer = request.META.get("HTTP_REFERER", "/history-admin/")
    if not tx_date or qty <= 0:
        return redirect(referer + ("&" if "?" in referer else "?") + "error=거래일과 수량을 올바르게 입력해 주세요.")
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE transaction_history SET tx_type = %s, tx_date = %s, qty = %s WHERE id = %s",
        (tx_type, tx_date, qty, record_id),
    )
    conn.commit()
    conn.close()
    from app.services.history_service import recalculate_current_stock
    recalculate_current_stock()
    return redirect(referer + ("&" if "?" in referer else "?") + "message=수정되었습니다.")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def history_admin_delete(request, record_id):
    if request.method != "POST":
        return redirect("history_admin")
    referer = request.META.get("HTTP_REFERER", "/history-admin/")
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM barcode_units WHERE inbound_tx_id = %s OR outbound_tx_id = %s", (record_id, record_id))
    cursor.execute("DELETE FROM transaction_history WHERE id = %s", (record_id,))
    conn.commit()
    conn.close()
    from app.services.history_service import recalculate_current_stock
    recalculate_current_stock()
    return redirect(referer + ("&" if "?" in referer else "?") + "message=삭제되었습니다.")


@login_required
@user_passes_test(can_access_admin_area)
def inventory_admin_page(request):
    q = request.GET.get("q", "")
    equipment = request.GET.get("equipment", "")
    reagent_type = request.GET.get("reagent_type", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "asc")
    if "part" in request.GET:
        part = request.GET.get("part", "")
    else:
        part = getattr(getattr(request.user, "profile", None), "part", "") or ""
    items = get_inventory_items(part=part, q=q, equipment=equipment, reagent_type=reagent_type, sort=sort, order=order)
    filter_options = get_inventory_filter_options()
    context = {
        "active_menu": "admin_panel",
        "items": items,
        "part": part,
        "q": q,
        "equipment": equipment,
        "reagent_type": reagent_type,
        "sort": sort,
        "order": order,
        "filter_options": filter_options,
        "part_map": get_part_map(get_current_schema()),
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
    }
    return render(request, "inventory_admin.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def inventory_admin_edit(request, item_id):
    from django.http import JsonResponse
    from datetime import date
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
    new_stock_str = request.POST.get("current_stock", "").strip()
    if new_stock_str == "" or not new_stock_str.lstrip("-").isdigit():
        return JsonResponse({"ok": False, "error": "재고 수량을 올바르게 입력해 주세요."}, status=400)
    new_stock = int(new_stock_str)

    created_by = request.user.get_full_name() or request.user.username
    created_by_empno = ""

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM inventory WHERE id = %s AND disposed_at IS NULL", (item_id,))
        row = cursor.fetchone()
        if not row:
            return JsonResponse({"ok": False, "error": "항목을 찾을 수 없습니다."}, status=404)
        item = dict(row)
        old_stock = int(item.get("current_stock", 0) or 0)
        delta = new_stock - old_stock
        if delta == 0:
            return JsonResponse({"ok": True})

        safety_stock = int(item.get("safety_stock", 0) or 0)
        required_qty = max(safety_stock - new_stock, 0)
        tx_type = "IN" if delta > 0 else "OUT"
        qty = abs(delta)
        tx_date = date.today().isoformat()

        cursor.execute(
            "UPDATE inventory SET current_stock = %s, required_qty = %s WHERE id = %s",
            (new_stock, required_qty, item_id),
        )
        cursor.execute(
            """
            INSERT INTO transaction_history (
                inventory_id, tx_type, qty, tx_date, note, remaining_stock,
                item_code, item_name, lot_no, part, unit,
                created_by, created_by_empno
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                item_id, tx_type, qty, tx_date, "관리자 재고 수정", new_stock,
                item.get("item_code", ""), item.get("item_name", ""),
                item.get("lot_no", ""), item.get("part", ""), item.get("unit", ""),
                created_by, created_by_empno,
            ),
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return JsonResponse({"ok": False, "error": str(e)}, status=500)
    finally:
        conn.close()

    return JsonResponse({"ok": True})


@login_required
def reagent_history_page(request):
    part = _get_part(request)
    q = request.GET.get("q", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    hazardous_grade = request.GET.get("hazardous_grade", "")
    is_new_lot = request.GET.get("is_new_lot", "")
    disposed = request.GET.get("disposed", "")
    show_form = request.GET.get("show_form", "")
    selected_item_id = request.GET.get("selected_item_id", "")
    selected_item_label = request.GET.get("selected_item_label", "")
    selected_ids = request.GET.get("selected_ids", "")
    _forced = "" if _is_dlab(request) else "ZZ"
    old_new_part = _forced or request.GET.get("old_new_part", "")
    old_new_mode = request.GET.get("old_new_mode", "")
    manage_part = _forced or request.GET.get("manage_part", "")
    manage_mode = request.GET.get("manage_mode", "")

    effective_part = manage_part or part
    effective_q = q
    effective_reagent_type = reagent_type
    effective_equipment = equipment
    effective_vendor = vendor
    effective_hazardous = hazardous
    effective_hazardous_grade = hazardous_grade
    effective_is_new_lot = is_new_lot
    effective_disposed = disposed
    effective_lot_status = "NEW" if manage_mode == "new" else ""

    from datetime import datetime

    items = get_reagent_history_items(
        part=effective_part,
        q=effective_q,
        reagent_type=effective_reagent_type,
        equipment=effective_equipment,
        vendor=effective_vendor,
        hazardous=effective_hazardous,
        hazardous_grade=effective_hazardous_grade,
        is_new_lot=effective_is_new_lot,
        disposed=effective_disposed,
        lot_status=effective_lot_status,
        sort=sort,
        order=order,
    )
    context = {
        "active_menu": "reagent_history",
        "items": items,
        "part_map": get_part_map(get_current_schema()),
        "part": part,
        "q": q,
        "sort": sort,
        "order": order,
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
        "reagent_type": reagent_type,
        "equipment": equipment,
        "vendor": vendor,
        "hazardous": hazardous,
        "hazardous_grade": hazardous_grade,
        "is_new_lot": is_new_lot,
        "disposed": disposed,
        "show_form": show_form,
        "selected_item_id": selected_item_id,
        "selected_item_label": selected_item_label,
        "selected_ids": selected_ids,
        "old_new_part": old_new_part,
        "old_new_mode": old_new_mode,
        "manage_part": manage_part,
        "manage_mode": manage_mode,
        "old_new_items": get_old_new_lot_items(part=old_new_part, only_new=(old_new_mode == "new")),
        "default_date": datetime.now().strftime("%Y-%m-%d"),
        "filter_options": get_reagent_history_filter_options(),
        "filtered_equipments": get_part_equipment_options(part),
        "filtered_vendors": get_all_vendor_options(),
    }
    return render(request, "reagent_history.html", context)


@login_required
def reagent_history_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    part = request.GET.get("part", "")
    q = request.GET.get("q", "")
    reagent_type = request.GET.get("reagent_type", "")
    equipment = request.GET.get("equipment", "")
    vendor = request.GET.get("vendor", "")
    hazardous = request.GET.get("hazardous", "")
    disposed = request.GET.get("disposed", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")

    items = get_reagent_history_items(
        part=part, q=q, reagent_type=reagent_type, equipment=equipment,
        vendor=vendor, hazardous=hazardous, disposed=disposed,
        sort=sort, order=order,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시약이력"

    headers = ["파트", "품목코드", "품목명", "Lot No", "유효기간", "규격", "단위",
               "시약종류", "장비", "업체", "현재재고", "안전재고", "유해화학물질",
               "개봉일", "병행사용시작일", "폐기일", "Lot 상태"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item.get("part_label", ""), item.get("item_code", ""), item.get("item_name", ""),
            item.get("lot_no") or "", item.get("expiry_date") or "", item.get("spec") or "",
            item.get("unit") or "", item.get("reagent_type") or "", item.get("equipment") or "",
            item.get("vendor") or "", item.get("current_stock", 0), item.get("safety_stock", 0),
            item.get("hazardous") or "", item.get("opened_at") or "",
            item.get("parallel_at") or "", item.get("disposed_at") or "",
            item.get("lot_status") or "",
        ])

    col_widths = [16, 14, 40, 16, 12, 16, 8, 14, 18, 18, 10, 10, 14, 12, 14, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    encoded = quote("시약이력.xlsx")
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@csrf_exempt
@login_required
def reagent_history_opened_at(request):
    if request.method != "POST":
        return redirect("reagent_history")

    item_id = int(request.POST.get("item_id", "0") or 0)
    opened_at = request.POST.get("opened_at", "")
    manage_part = request.POST.get("manage_part", "")
    manage_mode = request.POST.get("manage_mode", "")
    q = request.POST.get("q", "")
    sort = request.POST.get("sort", "")
    order = request.POST.get("order", "")

    ok, msg = update_opened_at(item_id=item_id, opened_at=opened_at)
    query = (
        f"?manage_part={quote_plus(manage_part)}&manage_mode={quote_plus(manage_mode)}"
        f"&q={quote_plus(q)}&sort={quote_plus(sort)}&order={quote_plus(order)}"
    )
    if ok:
        return redirect(f"/reagent-history/{query}&message={quote_plus('개봉 날짜가 등록되었습니다.')}")
    return redirect(f"/reagent-history/{query}&show_form=opened&error={quote_plus(msg)}")


@csrf_exempt
@login_required
def reagent_history_parallel_at(request):
    if request.method != "POST":
        return redirect("reagent_history")

    item_id = int(request.POST.get("item_id", "0") or 0)
    parallel_at = request.POST.get("parallel_at", "")
    manage_part = request.POST.get("manage_part", "")
    manage_mode = request.POST.get("manage_mode", "")
    q = request.POST.get("q", "")
    sort = request.POST.get("sort", "")
    order = request.POST.get("order", "")

    ok, msg = update_parallel_at(item_id=item_id, parallel_at=parallel_at)
    query = (
        f"?manage_part={quote_plus(manage_part)}&manage_mode={quote_plus(manage_mode)}"
        f"&q={quote_plus(q)}&sort={quote_plus(sort)}&order={quote_plus(order)}"
    )
    if ok:
        return redirect(f"/reagent-history/{query}&message={quote_plus('Parallel 날짜가 등록되었습니다.')}")
    return redirect(f"/reagent-history/{query}&show_form=parallel&error={quote_plus(msg)}")


@csrf_exempt
@login_required
def reagent_history_old_new_lot_save(request):
    if request.method != "POST":
        return redirect("reagent_history")

    part = request.POST.get("part", "")
    visible_item_ids = [int(item_id) for item_id in request.POST.getlist("visible_item_ids") if str(item_id).strip()]
    new_lot_item_ids = [int(item_id) for item_id in request.POST.getlist("new_lot_item_ids") if str(item_id).strip()]

    ok, msg = save_old_new_lot_selection(
        part=part,
        visible_item_ids=visible_item_ids,
        new_lot_item_ids=new_lot_item_ids,
    )
    target = f"/reagent-history/?old_new_part={quote_plus(part)}"
    if ok:
        return redirect(f"{target}&message={quote_plus(msg)}")
    return redirect(f"{target}&error={quote_plus(msg)}")


# ──────────────────────────────────────────────────────────
# Raw DB 관리
# ──────────────────────────────────────────────────────────

@login_required
@user_passes_test(can_access_admin_area)
def rawdb_admin_page(request):
    part = _get_part(request)
    q = request.GET.get("q", "")
    hazardous_grade = request.GET.get("hazardous_grade", "")
    not_in_master = bool(request.GET.get("not_in_master", ""))
    in_master = request.GET.get("in_master", "")
    sort = request.GET.get("sort", "")
    order = request.GET.get("order", "")
    show_form = request.GET.get("show_form", "")
    edit_id = request.GET.get("edit_id", "")

    current_schema = get_current_schema()
    schema_part_map = get_part_map(current_schema)

    context = {
        "active_menu": "admin_panel",
        "current_schema": current_schema,
        "items": get_rawdb_items(part=part, q=q, hazardous_grade=hazardous_grade,
                                 not_in_master=not_in_master, in_master=in_master, sort=sort, order=order,
                                 schema=current_schema),
        "part_map": schema_part_map,
        "filter_options": get_rawdb_filter_options(),
        "part": part,
        "q": q,
        "hazardous_grade": hazardous_grade,
        "not_in_master": not_in_master,
        "in_master": in_master,
        "sort": sort,
        "order": order,
        "show_form": show_form,
        "edit_item": get_rawdb_item_by_id(int(edit_id)) if edit_id else None,
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
    }
    return render(request, "rawdb_admin.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_create(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/")
    create_rawdb_item(
        part=request.POST.get("part", ""),
        hazardous=request.POST.get("hazardous", "N"),
        hazardous_grade=request.POST.get("hazardous_grade", ""),
        item_code=request.POST.get("item_code", ""),
        item_name=request.POST.get("item_name", ""),
        spec=request.POST.get("spec", ""),
        unit=request.POST.get("unit", ""),
        lot_no=request.POST.get("lot_no", ""),
        expiry_date=request.POST.get("expiry_date", ""),
        reagent_type=request.POST.get("reagent_type", ""),
        vendor=request.POST.get("vendor", ""),
        equipment=request.POST.get("equipment", ""),
        schema=get_current_schema(),
    )
    return redirect("/rawdb-admin/?message=항목이 등록되었습니다.")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_edit_submit(request, item_id: int):
    if request.method != "POST":
        return redirect("/rawdb-admin/")
    update_rawdb_item(
        item_id=item_id,
        part=request.POST.get("part", ""),
        hazardous=request.POST.get("hazardous", "N"),
        hazardous_grade=request.POST.get("hazardous_grade", ""),
        item_code=request.POST.get("item_code", ""),
        item_name=request.POST.get("item_name", ""),
        spec=request.POST.get("spec", ""),
        unit=request.POST.get("unit", ""),
        lot_no=request.POST.get("lot_no", ""),
        expiry_date=request.POST.get("expiry_date", ""),
        reagent_type=request.POST.get("reagent_type", ""),
        vendor=request.POST.get("vendor", ""),
        equipment=request.POST.get("equipment", ""),
    )
    return redirect("/rawdb-admin/?message=항목이 수정되었습니다.")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_delete(request, item_id: int):
    if request.method != "POST":
        return redirect("/rawdb-admin/")
    delete_rawdb_item(item_id)
    return redirect("/rawdb-admin/?message=항목이 삭제되었습니다.")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_bulk_delete(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/")
    item_ids = [int(i) for i in request.POST.getlist("item_ids") if str(i).strip()]
    bulk_delete_rawdb_items(item_ids)
    return redirect(f"/rawdb-admin/?message={len(item_ids)}개 항목이 삭제되었습니다.")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_upload_preview_view(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/?not_in_master=1")

    item_ids = request.POST.getlist("item_ids")
    if not item_ids:
        return redirect("/rawdb-admin/?not_in_master=1&error=업로드할 항목을 선택해 주세요.")

    preview_items = [get_rawdb_item_by_id(int(iid)) for iid in item_ids]
    preview_items = [item for item in preview_items if item]

    part = request.POST.get("part", "")
    q = request.POST.get("q", "")
    hazardous_grade = request.POST.get("hazardous_grade", "")
    sort = request.POST.get("sort", "")
    order = request.POST.get("order", "")

    schema = get_current_schema()
    vendor_equipment = get_master_vendor_equipment_options(schema) if schema else {"vendors": [], "equipments": []}

    context = {
        "active_menu": "admin_panel",
        "items": get_rawdb_items(part=part, q=q, hazardous_grade=hazardous_grade,
                                 not_in_master=True, sort=sort, order=order),
        "part_map": get_part_map(schema),
        "filter_options": get_rawdb_filter_options(),
        "part": part,
        "q": q,
        "hazardous_grade": hazardous_grade,
        "not_in_master": True,
        "sort": sort,
        "order": order,
        "show_form": "",
        "edit_item": None,
        "message": "",
        "error": "",
        "preview_mode": True,
        "preview_items": preview_items,
        "vendor_options": vendor_equipment["vendors"],
        "equipment_options": vendor_equipment["equipments"],
    }
    return render(request, "rawdb_admin.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def rawdb_upload_to_master_view(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/?not_in_master=1")
    schema = get_current_schema()
    if not schema:
        return redirect("/rawdb-admin/?not_in_master=1&error=부서가 설정되지 않았습니다.")

    item_ids = request.POST.getlist("item_ids")
    if not item_ids:
        return redirect("/rawdb-admin/?not_in_master=1&error=업로드할 항목을 선택해 주세요.")

    items_data = [
        {
            "item_id": int(iid),
            "lot_no": request.POST.get(f"lot_no_{iid}", ""),
            "expiry_date": request.POST.get(f"expiry_date_{iid}", ""),
            "reagent_type": request.POST.get(f"reagent_type_{iid}", ""),
            "vendor": request.POST.get(f"vendor_{iid}", ""),
            "equipment": request.POST.get(f"equipment_{iid}", ""),
        }
        for iid in item_ids
    ]

    success, errors = upload_rawdb_items_to_master(schema, items_data)
    msg = f"{success}건이 시약 마스터에 등록되었습니다."
    if errors:
        err_msg = "; ".join(errors[:3]) + ("..." if len(errors) > 3 else "")
        return redirect(f"/rawdb-admin/?not_in_master=1&message={msg}&error={err_msg}")
    return redirect(f"/rawdb-admin/?not_in_master=1&message={msg}")


@login_required
@user_passes_test(can_access_admin_area)
def rawdb_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from urllib.parse import quote

    part          = request.GET.get("part", "")
    q             = request.GET.get("q", "")
    hazardous_grade = request.GET.get("hazardous_grade", "")
    not_in_master = bool(request.GET.get("not_in_master", ""))
    in_master     = request.GET.get("in_master", "")
    sort          = request.GET.get("sort", "")
    order         = request.GET.get("order", "")

    items = get_rawdb_items(
        part=part, q=q, hazardous_grade=hazardous_grade,
        not_in_master=not_in_master, in_master=in_master,
        sort=sort, order=order,
    )

    from app.utils.constants import REAGENT_TYPE_MAP
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RawDB"

    headers = ["파트", "유해물질", "등급", "품목코드", "품목명",
               "규격", "단위", "Lot No", "유효기간", "시약 구분", "업체", "단가", "장비", "마스터 등록"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        reagent_raw = str(item.get("reagent_type") or "").strip()
        ws.append([
            item.get("part", ""),
            item.get("hazardous", ""),
            item.get("hazardous_grade", ""),
            item.get("item_code", ""),
            item.get("item_name", ""),
            item.get("spec", ""),
            item.get("unit", ""),
            item.get("lot_no", ""),
            item.get("expiry_date", ""),
            REAGENT_TYPE_MAP.get(reagent_raw, reagent_raw),
            item.get("vendor", ""),
            item.get("unit_price"),
            item.get("equipment", ""),
            item.get("in_master", ""),
        ])

    col_widths = [12, 10, 14, 16, 40, 16, 8, 16, 12, 14, 18, 12, 18, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote("RawDB.xlsx")
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded}"
    return response


@user_passes_test(lambda u: u.is_superuser)
def rawdb_sync_view(request):
    if request.method != "POST":
        return redirect("/admin-panel/")
    result = sync_rawdb_to_inventory()
    total = sum(result.values())
    detail = ", ".join(f"{s}:{n}" for s, n in result.items())
    return redirect(f"/admin-panel/?message=Raw DB 동기화 완료: 총 {total}건 업데이트 ({detail})")


# ── 업체 목록 관리 ────────────────────────────────────────────────

@login_required
@user_passes_test(can_access_admin_area)
def vendor_list_page(request):
    edit_id = request.GET.get("edit_id", "")
    context = {
        "active_menu": "admin_panel",
        "list_type": "vendor",
        "title": "업체 목록 관리",
        "items": get_vendor_list(),
        "edit_id": int(edit_id) if edit_id else None,
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
        "back_url": "/rawdb-admin/",
        "create_url": "/rawdb-admin/vendor-list/create/",
        "list_url": "/rawdb-admin/vendor-list/",
    }
    return render(request, "vendor_equipment_list.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def vendor_list_create(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/vendor-list/")
    name = request.POST.get("name", "").strip()
    if not name:
        return redirect("/rawdb-admin/vendor-list/?error=업체명을 입력해주세요.")
    create_vendor(name)
    return redirect(f"/rawdb-admin/vendor-list/?message={name} 등록 완료")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def vendor_list_edit(request, item_id: int):
    if request.method != "POST":
        return redirect("/rawdb-admin/vendor-list/")
    name = request.POST.get("name", "").strip()
    if not name:
        return redirect("/rawdb-admin/vendor-list/?error=업체명을 입력해주세요.")
    update_vendor(item_id, name)
    return redirect(f"/rawdb-admin/vendor-list/?message={name} 수정 완료")


# ── 장비 목록 관리 ────────────────────────────────────────────────

@login_required
@user_passes_test(can_access_admin_area)
def equipment_list_page(request):
    edit_id = request.GET.get("edit_id", "")
    filter_part = request.GET.get("filter_part", "")
    context = {
        "active_menu": "admin_panel",
        "list_type": "equipment",
        "title": "장비 목록 관리",
        "items": get_equipment_list(part=filter_part),
        "part_map": get_part_map(get_current_schema()),
        "filter_part": filter_part,
        "edit_id": int(edit_id) if edit_id else None,
        "message": request.GET.get("message", ""),
        "error": request.GET.get("error", ""),
        "back_url": "/rawdb-admin/",
        "create_url": "/rawdb-admin/equipment-list/create/",
        "list_url": "/rawdb-admin/equipment-list/",
    }
    return render(request, "vendor_equipment_list.html", context)


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def equipment_list_create(request):
    if request.method != "POST":
        return redirect("/rawdb-admin/equipment-list/")
    name = request.POST.get("name", "").strip()
    part = request.POST.get("part", "").strip()
    if not name:
        return redirect("/rawdb-admin/equipment-list/?error=장비명을 입력해주세요.")
    create_equipment(name, part)
    return redirect(f"/rawdb-admin/equipment-list/?message={name} 등록 완료")


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def equipment_list_edit(request, item_id: int):
    if request.method != "POST":
        return redirect("/rawdb-admin/equipment-list/")
    name = request.POST.get("name", "").strip()
    part = request.POST.get("part", "").strip()
    if not name:
        return redirect("/rawdb-admin/equipment-list/?error=장비명을 입력해주세요.")
    update_equipment(item_id, name, part)
    return redirect(f"/rawdb-admin/equipment-list/?message={name} 수정 완료")


@login_required
@user_passes_test(can_access_admin_area)
def disposal_admin_page(request):
    part = request.GET.get("part", "")
    q = request.GET.get("q", "")
    message = request.GET.get("message", "")
    error = request.GET.get("error", "")
    items = get_disposed_items(part=part, q=q)
    return render(request, "disposal_admin.html", {
        "active_menu": "admin_panel",
        "items": items,
        "part": part,
        "q": q,
        "message": message,
        "error": error,
        "part_map": get_part_map(get_current_schema()),
    })


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def disposal_admin_cancel(request):
    if request.method != "POST":
        return redirect("disposal_admin")
    ids = request.POST.getlist("item_ids")
    if not ids:
        return redirect("/disposal-admin/?error=선택된 항목이 없습니다.")
    for item_id in ids:
        new_expiry = request.POST.get(f"new_expiry_{item_id}", "").strip()
        cancel_dispose(int(item_id), new_expiry=new_expiry or None)
    return redirect(f"/disposal-admin/?message={len(ids)}개 항목의 폐기가 취소되었습니다.")


@login_required
@user_passes_test(can_access_admin_area)
def preset_admin_page(request):
    part = request.GET.get("part", "")
    message = request.GET.get("message", "")
    error = request.GET.get("error", "")
    presets = get_presets(part=part)
    return render(request, "preset_admin.html", {
        "active_menu": "admin_panel",
        "presets": presets,
        "part": part,
        "message": message,
        "error": error,
        "part_map": get_part_map(get_current_schema()),
    })


@login_required
@user_passes_test(can_access_admin_area)
def preset_form_page(request, preset_id=None):
    part_map = get_part_map(get_current_schema())
    preset = get_preset(preset_id) if preset_id else None
    selected_part = request.GET.get("part", preset["part"] if preset else "")
    item_codes = get_part_item_codes(selected_part) if selected_part else []
    preset_codes = {item["item_code"] for item in preset["items"]} if preset else set()

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        part = request.POST.get("part", "").strip()
        codes = request.POST.getlist("item_codes")
        if not name or not part or not codes:
            return render(request, "preset_form.html", {
                "active_menu": "admin_panel",
                "preset": preset,
                "part_map": part_map,
                "selected_part": part,
                "item_codes": get_part_item_codes(part),
                "preset_codes": set(codes),
                "error": "약속명, 파트, 품목을 모두 선택해 주세요.",
            })
        if preset_id:
            update_preset(preset_id, name, codes)
            return redirect(f"/preset-admin/?part={part}&message={name} 수정 완료")
        else:
            create_preset(name, part, codes)
            return redirect(f"/preset-admin/?part={part}&message={name} 등록 완료")

    return render(request, "preset_form.html", {
        "active_menu": "admin_panel",
        "preset": preset,
        "part_map": part_map,
        "selected_part": selected_part,
        "item_codes": item_codes,
        "preset_codes": preset_codes,
    })


@csrf_exempt
@login_required
@user_passes_test(can_access_admin_area)
def preset_delete(request, preset_id):
    if request.method == "POST":
        preset = get_preset(preset_id)
        part = preset["part"] if preset else ""
        delete_preset(preset_id)
        return redirect(f"/preset-admin/?part={part}&message=삭제 완료")
    return redirect("/preset-admin/")


@login_required
def preset_cart_api(request, preset_id):
    items = get_preset_cart_items(preset_id)
    from app.utils.constants import REAGENT_TYPE_MAP
    result = []
    for item in items:
        reagent_type_code = str(item.get("reagent_type", "")).strip()
        result.append({
            "id": item["id"],
            "item_code": item["item_code"],
            "item_name": item["item_name"],
            "lot_no": item["lot_no"] or "",
            "expiry_date": str(item.get("expiry_date") or ""),
            "reagent_type": REAGENT_TYPE_MAP.get(reagent_type_code, reagent_type_code),
            "vendor": item.get("vendor") or "",
            "part": item.get("part") or "",
            "unit": item.get("unit") or "",
            "spec": item.get("spec") or "",
        })
    return JsonResponse({"items": result})


@login_required
def help_page(request):
    return render(request, "help.html", {"active_menu": "help"})
